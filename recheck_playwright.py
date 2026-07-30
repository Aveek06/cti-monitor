"""
Headless-browser re-check for sites that failed the plain-requests feed check.

Why this exists: `requests` doesn't execute JavaScript and doesn't look like a
real browser to bot-protection systems. This script uses Playwright (real
Chromium) to load each failed page properly, then:
  1. Checks the actual HTTP response status Chromium got (some 403s from
     `requests` disappear when a real browser loads the page).
  2. Searches the fully-rendered HTML (after JS execution) for a
     <link rel="alternate" type="application/rss+xml|atom+xml"> tag.
  3. As a fallback, scans all rendered <a> and <link> hrefs for anything
     containing rss/feed/atom/.xml as a rough heuristic.

Usage:
    pip install playwright
    playwright install chromium      # one-time browser download
    python recheck_playwright.py CTI_Feed_Check_Final.xlsx CTI_Feed_Check_Playwright.xlsx
"""

import re
import sys
import time
import openpyxl
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

NAV_TIMEOUT_MS = 25000
FEED_HREF_PATTERN = re.compile(r"(rss|atom|feed)", re.IGNORECASE)
XML_EXT_PATTERN = re.compile(r"\.xml($|\?)", re.IGNORECASE)


def check_with_browser(page, url):
    result = {"feed_found": False, "feed_url": None, "method": None, "notes": ""}
    try:
        resp = page.goto(url, timeout=NAV_TIMEOUT_MS, wait_until="domcontentloaded")
        try:
            page.wait_for_load_state("networkidle", timeout=8000)
        except PWTimeout:
            pass  # some sites never go fully idle (ads, trackers) - fine, proceed anyway

        status = resp.status if resp else None
        if status and status >= 400:
            result["notes"] = f"browser also got status {status}"
            return result

        # 1. Proper <link rel="alternate"> tag search
        link_handle = page.query_selector(
            'link[rel="alternate"][type*="rss"], link[rel="alternate"][type*="atom"]'
        )
        if link_handle:
            href = link_handle.get_attribute("href")
            if href:
                feed_url = page.evaluate("(h) => new URL(h, document.baseURI).href", href)
                result.update(feed_found=True, feed_url=feed_url, method="browser-link-tag")
                return result

        # 2. Fallback: scan all hrefs on the rendered page for feed-like links
        hrefs = page.eval_on_selector_all("a[href], link[href]", "els => els.map(e => e.href)")
        candidates = [h for h in hrefs if FEED_HREF_PATTERN.search(h) or XML_EXT_PATTERN.search(h)]
        # de-prioritize obvious false positives like social share/feed icons pointing to non-xml pages
        candidates = [h for h in candidates if not any(x in h.lower() for x in
                      ["facebook", "twitter", "linkedin", "feedback", "feedly.com/i/subscription"])]
        if candidates:
            result.update(feed_found=True, feed_url=candidates[0], method="browser-href-heuristic",
                           notes=f"{len(candidates)} candidate link(s) found, took first")
            return result

        result["notes"] = "page loaded fine, no feed reference found even after JS render"
        return result

    except PWTimeout:
        result["notes"] = "browser navigation timeout"
        return result
    except Exception as e:
        result["notes"] = f"browser error: {e.__class__.__name__}: {str(e)[:120]}"
        return result


def main(input_path, output_path):
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    header = ["Blog Name", "Original URL", "Feed Found", "Feed URL", "Detection Method", "Notes"]

    successes, to_recheck = [], []
    for r in rows:
        d = dict(zip(header, r))
        (successes if d["Feed Found"] == "Yes" else to_recheck).append(d)

    print(f"Re-checking {len(to_recheck)} sites with a real headless browser...")

    rechecked = []
    with sync_playwright() as p:
        # --disable-http2 works around ERR_HTTP2_PROTOCOL_ERROR seen on some
        # sites (McAfee, Trellix, Sophos, Secureworks) that Chromium's HTTP2
        # implementation doesn't get along with.
        browser = p.chromium.launch(headless=True, args=["--disable-http2"])
        context = browser.new_context(
            user_agent=("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"),
            viewport={"width": 1366, "height": 900},
            ignore_https_errors=True,  # works around Phylum's CERT_COMMON_NAME_INVALID
        )
        page = context.new_page()

        for i, d in enumerate(to_recheck, 1):
            url = d["Original URL"].split("\n")[0].strip()  # first URL if multiple stacked
            res = check_with_browser(page, url)
            rechecked.append({
                "Blog Name": d["Blog Name"],
                "Original URL": d["Original URL"],
                "Feed Found": "Yes" if res["feed_found"] else "No",
                "Feed URL": res.get("feed_url") or "",
                "Detection Method": res.get("method") or "",
                "Notes": res.get("notes") or "",
            })
            status = "RECOVERED" if res["feed_found"] else "still no feed"
            print(f"[{i}/{len(to_recheck)}] {d['Blog Name']}: {status}")
            time.sleep(1)  # be a decent citizen, don't hammer sites back-to-back

        browser.close()

    recovered = sum(1 for r in rechecked if r["Feed Found"] == "Yes")
    print(f"\nRecovered {recovered}/{len(to_recheck)} additional sites via headless browser.")

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Playwright Recheck Results"
    out_ws.append(header)
    for d in successes:
        out_ws.append([d[h] for h in header])
    for d in rechecked:
        out_ws.append([d[h] for h in header])
    out_wb.save(output_path)

    total_yes = len(successes) + recovered
    total = len(successes) + len(rechecked)
    print(f"\nFinal tally: {total_yes}/{total} sites have a detectable feed.")
    print(f"Genuinely feed-less (need scraping decision): {total - total_yes}")
    print(f"Saved to {output_path}")


if __name__ == "__main__":
    inp = sys.argv[1] if len(sys.argv) > 1 else "CTI_Feed_Check_Final.xlsx"
    out = sys.argv[2] if len(sys.argv) > 2 else "CTI_Feed_Check_Playwright.xlsx"
    main(inp, out)
