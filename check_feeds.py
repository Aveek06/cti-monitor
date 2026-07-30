"""
CTI Source Feed Checker
------------------------
Reads CTI_Source_List.xlsx (columns: Blog Name, Blog Site) and, for each
site, tries to find a working RSS/Atom feed. Writes results to
CTI_Feed_Check_Results.xlsx with columns:

    Blog Name | Blog Site | Feed Found | Feed URL | Detection Method | Notes

Run this where you have real internet access (e.g. via Claude Code, or
locally with `python check_feeds.py CTI_Source_List.xlsx`).

Requires: requests, beautifulsoup4, openpyxl, lxml
    pip install requests beautifulsoup4 openpyxl lxml
"""

import sys
import time
import concurrent.futures as cf
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
import openpyxl

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/124.0.0.0 Safari/537.36"
}
TIMEOUT = 12

# Common feed path patterns to try relative to the site's origin.
COMMON_FEED_PATHS = [
    "/feed/",
    "/feed",
    "/rss/",
    "/rss.xml",
    "/rss",
    "/atom.xml",
    "/index.xml",       # Hugo default
    "/blog/feed/",
    "/blog/rss.xml",
    "/feeds/posts/default",   # Blogger
]

FEED_CONTENT_MARKERS = ("<rss", "<feed", "application/rss+xml", "application/atom+xml")


def looks_like_feed(resp: requests.Response) -> bool:
    ctype = resp.headers.get("Content-Type", "").lower()
    if "xml" in ctype or "rss" in ctype or "atom" in ctype:
        return True
    head = resp.text[:500].lower() if resp.text else ""
    return any(marker in head for marker in FEED_CONTENT_MARKERS)


def try_common_paths(origin: str):
    for path in COMMON_FEED_PATHS:
        url = urljoin(origin, path)
        try:
            r = requests.get(url, headers=HEADERS, timeout=TIMEOUT, allow_redirects=True)
            if r.status_code == 200 and looks_like_feed(r):
                return url, "common-path"
        except requests.RequestException:
            continue
    return None, None


def try_autodiscovery(page_url: str):
    try:
        r = requests.get(page_url, headers=HEADERS, timeout=TIMEOUT, allow_redirects=True)
        if r.status_code != 200:
            return None, None, f"page returned {r.status_code}"
        soup = BeautifulSoup(r.text, "lxml")
        link = soup.find("link", rel=lambda v: v and "alternate" in v,
                          type=lambda t: t and ("rss" in t or "atom" in t))
        if link and link.get("href"):
            feed_url = urljoin(page_url, link["href"])
            # verify it actually resolves
            try:
                fr = requests.get(feed_url, headers=HEADERS, timeout=TIMEOUT)
                if fr.status_code == 200 and looks_like_feed(fr):
                    return feed_url, "autodiscovery", None
            except requests.RequestException:
                pass
            return feed_url, "autodiscovery-unverified", None
        return None, None, "no feed link tag found"
    except requests.RequestException as e:
        return None, None, f"error: {e.__class__.__name__}"


def check_one_url(page_url: str):
    page_url = page_url.strip()
    if not page_url:
        return {"feed_found": False, "feed_url": None, "method": None, "notes": "blank URL"}

    parsed = urlparse(page_url)
    origin = f"{parsed.scheme}://{parsed.netloc}"

    feed_url, method = try_common_paths(origin)
    if feed_url:
        return {"feed_found": True, "feed_url": feed_url, "method": method, "notes": ""}

    feed_url, method, note = try_autodiscovery(page_url)
    if feed_url:
        return {"feed_found": True, "feed_url": feed_url, "method": method, "notes": note or ""}

    return {"feed_found": False, "feed_url": None, "method": None, "notes": note or "no feed found"}


def check_site(blog_name: str, blog_site_cell: str):
    # Some cells contain multiple URLs stacked with newlines
    urls = [u.strip() for u in blog_site_cell.split("\n") if u.strip()]
    results = []
    for url in urls:
        r = check_one_url(url)
        r["blog_name"] = blog_name
        r["original_url"] = url
        results.append(r)
        if r["feed_found"]:
            break  # good enough, stop checking other URLs for this row
    return results


def main(input_path, output_path):
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header

    all_results = []
    with cf.ThreadPoolExecutor(max_workers=10) as ex:
        futures = {ex.submit(check_site, name, site): (name, site) for name, site in rows if name}
        for i, fut in enumerate(cf.as_completed(futures), 1):
            name, site = futures[fut]
            try:
                res = fut.result()
            except Exception as e:
                res = [{"blog_name": name, "original_url": site, "feed_found": False,
                        "feed_url": None, "method": None, "notes": f"crashed: {e}"}]
            all_results.extend(res)
            print(f"[{i}/{len(futures)}] {name}: "
                  f"{'FEED FOUND' if res[0]['feed_found'] else 'no feed'}")

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Feed Check Results"
    out_ws.append(["Blog Name", "Original URL", "Feed Found", "Feed URL", "Detection Method", "Notes"])
    for r in all_results:
        out_ws.append([
            r["blog_name"], r["original_url"], "Yes" if r["feed_found"] else "No",
            r.get("feed_url") or "", r.get("method") or "", r.get("notes") or "",
        ])
    out_wb.save(output_path)
    found = sum(1 for r in all_results if r["feed_found"])
    print(f"\nDone. {found}/{len(all_results)} sources have a detectable feed.")
    print(f"Results saved to {output_path}")


if __name__ == "__main__":
    input_file = sys.argv[1] if len(sys.argv) > 1 else "CTI_Source_List.xlsx"
    output_file = sys.argv[2] if len(sys.argv) > 2 else "CTI_Feed_Check_Results.xlsx"
    main(input_file, output_file)
