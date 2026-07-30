"""
Re-check failures from a previous feed-check run, then merge with the
original results into one final file.

Usage:
    python recheck_failures.py CTI_Feed_Check_Results.xlsx CTI_Feed_Check_Final.xlsx

Requires check_feeds.py to be in the same folder (imports check_one_url from it).
"""

import sys
import concurrent.futures as cf
import openpyxl

from check_feeds import check_one_url


def main(prev_results_path, final_output_path):
    wb = openpyxl.load_workbook(prev_results_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    header = ["Blog Name", "Original URL", "Feed Found", "Feed URL", "Detection Method", "Notes"]

    # Split into already-successful rows and failed rows to re-check
    successes = []
    failures = []
    for r in rows:
        d = dict(zip(header, r))
        if d["Feed Found"] == "Yes":
            successes.append(d)
        else:
            failures.append(d)

    print(f"Re-checking {len(failures)} previously failed sites...")

    rechecked = []
    with cf.ThreadPoolExecutor(max_workers=8) as ex:
        futures = {ex.submit(check_one_url, d["Original URL"]): d for d in failures}
        for i, fut in enumerate(cf.as_completed(futures), 1):
            d = futures[fut]
            try:
                res = fut.result()
            except Exception as e:
                res = {"feed_found": False, "feed_url": None, "method": None,
                       "notes": f"crashed on retry: {e}"}
            rechecked.append({
                "Blog Name": d["Blog Name"],
                "Original URL": d["Original URL"],
                "Feed Found": "Yes" if res["feed_found"] else "No",
                "Feed URL": res.get("feed_url") or "",
                "Detection Method": res.get("method") or "",
                "Notes": res.get("notes") or "",
            })
            status = "RECOVERED" if res["feed_found"] else "still no feed"
            print(f"[{i}/{len(failures)}] {d['Blog Name']}: {status}")

    recovered_count = sum(1 for r in rechecked if r["Feed Found"] == "Yes")
    print(f"\nRecovered {recovered_count}/{len(failures)} previously-failed sites.")

    # Merge: original successes + rechecked results, write final file
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Final Feed Results"
    out_ws.append(header)
    for d in successes:
        out_ws.append([d[h] for h in header])
    for d in rechecked:
        out_ws.append([d[h] for h in header])
    out_wb.save(final_output_path)

    total_yes = len(successes) + recovered_count
    total = len(successes) + len(rechecked)
    print(f"\nFinal tally: {total_yes}/{total} sites have a detectable feed.")
    print(f"Saved to {final_output_path}")


if __name__ == "__main__":
    prev = sys.argv[1] if len(sys.argv) > 1 else "CTI_Feed_Check_Results.xlsx"
    out = sys.argv[2] if len(sys.argv) > 2 else "CTI_Feed_Check_Final.xlsx"
    main(prev, out)
